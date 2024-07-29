import openpyxl


#in order to write same date:
# file="C:\\Users\\svankada\\Work\\Satish\\Personal\\Python\\V Satheesh Kumar\\Selenium-Python\\Scren shots\\SES-14 Data driven testing\\Empty.xlsx"
# workbook=openpyxl.load_workbook(file)
# #sheet=workbook["Sheet1"] #if you have multiple sheets, select by sheet name.
# sheet=workbook.active #if only one page availble.
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="welcome"
#
# workbook.save(file)

#multiple data:
#file==>workbook==>sheet==> rows==>cells

file="C:\\Users\\svankada\\Work\\Satish\\Personal\\Python\\V Satheesh Kumar\\Selenium-Python\\Scren shots\\SES-14 Data driven testing\\Empty_data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active
sheet.cell(1,1).value="Name"
sheet.cell(1,2).value="Marks"
sheet.cell(1,3).value="Class"


sheet.cell(2,1).value="sathesh"
sheet.cell(2,2).value=35
sheet.cell(2,3).value="PP1"

sheet.cell(3,1).value="harika"
sheet.cell(3,2).value=45
sheet.cell(3,3).value="PP1"

sheet.cell(4,1).value="joshvik"
sheet.cell(4,2).value=40
sheet.cell(4,3).value="PP1"

workbook.save(file)




