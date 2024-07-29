import openpyxl

#file==>workbook==>sheet==> rows==>cells
file="C:\\Users\\svankada\\Work\\Satish\\Personal\\Python\\V Satheesh Kumar\\Selenium-Python\\Scren shots\\SES-14 Data driven testing\\Sample.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row #19
columns=sheet.max_column #3

print(rows,columns)
#read the all columns and row from excel sheet.

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end="                  ")
    print()


